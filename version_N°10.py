#!/usr/bin/env python
# coding: utf-8

# In[ ]:


# 1 - # LIBRERÍAS NECESARIAS 

import pandas as pd
import numpy as np
import openpyxl 
import os
import sys
import shutil
import re # para expresiones regulares
import warnings
import traceback

import xlwings as xw

from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta

import tkinter as tk
from tkinter import Tk, Label, Listbox, Button, messagebox, Frame, font

# Filtrar todos los FutureWarnings para evitarlas en mi terminal al compilar
warnings.filterwarnings("ignore", category=FutureWarning)
warnings.filterwarnings("ignore", category=DeprecationWarning)



# In[ ]:


# 2 - # OBTIENE EL DIRECTORIO ACTUAL 

script_dir = os.getcwd()
script_dir = f"{script_dir}"  #" #AGREGAR /_internal"
script_dir = script_dir.replace('\\', "/")

print("### Buscando Ruta... ") 
print("")
print("Ruta Actual: ", script_dir)


# In[ ]:


# 3 - # INTERFAZ PARA SELECCIONAR OBRA

def INTERFAZ_SELECTOR_OBRA(ruta, obra):
    
    
    def seleccionar_obra():
        nonlocal ruta
        nonlocal obra
        seleccion = listbox.curselection()
        if seleccion:
            carpeta_elegida = carpetas_obras[int(seleccion[0])]
            ruta_completa = os.path.join(ruta_obras, carpeta_elegida)
            ruta = ruta_completa
            obra = carpeta_elegida
            
            if carpeta_elegida: 
                ventana.destroy()
                

    # Crear la ventana principal
    ventana = Tk()
    ventana.title("Selección de Obra")
    
    # Cambiar el tamaño de la fuente del título de la ventana
    font_titulo = font.Font(family='Calibri', size=15)  # Puedes ajustar la fuente y el tamaño aquí
    ventana.option_add("*Font", font_titulo)
    
    ventana.geometry("300x400")
    
    frame = Frame(ventana)
    frame.pack(padx=10, pady=5)
    
    # Ruta de la carpeta "obras"
    ruta_obras = f"{script_dir}/Obras" #/"agragar: /Estudio Dumont"
    
    if not os.path.exists(ruta_obras):
        tk.Label(ventana, text="La carpeta 'Obras' no existe en la ruta especificada.").pack()
        ventana.destroy()  # Cierra la ventana de tkinter
        return None
    
    # Obtener la lista de carpetas dentro de "obras"
    carpetas_obras = [nombre for nombre in os.listdir(ruta_obras) if os.path.isdir(os.path.join(ruta_obras, nombre))]

    # Mostrar las opciones al usuario en una lista
    label_font = font.Font(size=18) 
    tk.Label(frame, text="OBRAS DISPONIBLES:").pack()
    listbox_font = font.Font(size=18, font="calibri", bd=3)
    listbox = Listbox(frame, width=33, bd=3, relief='flat', highlightthickness=1)
    listbox.pack()

    for i, carpeta in enumerate(carpetas_obras, 1):
        listbox.insert(i, f"{i}. {carpeta}")

    # Botón para seleccionar la obra
    tk.Button(ventana, text="Seleccionar Obra", command=seleccionar_obra).pack()
    ventana.mainloop()  # Inicia el bucle principal de tkinter
    
    
    return ruta, obra

# Inicia la aplicación
# Función para completar la ruta
ruta = None
obra = None
ruta, nombre_carpeta = INTERFAZ_SELECTOR_OBRA(ruta, obra)

print(" PRIMERA PARTE: LECTURA")

# LEO Y CREO DATAFRAMES
# ABRIR PRESUPUESTO Y GASTOS CON OPENPYXL
print("Archivos en Carpeta:")
archivos_en_carpeta = [archivo for archivo in os.listdir(ruta) if os.path.isfile(os.path.join(ruta, archivo))]
for faz, bar in enumerate(archivos_en_carpeta):
    print(faz, bar)
print("")  

ruta_presupuesto = 0
ruta_gastos = 0
ruta_variables = 0
    
if ruta:
    
    #RUTAS 
    ruta_presupuesto = f"{ruta}/PRESUPUESTO GENERAL - {nombre_carpeta}.xlsx"
    ruta_gastos = f"{ruta}/GASTOS - {nombre_carpeta}.xlsx"
    ruta_variables = f"{script_dir}/VARIABLES.xlsx"  #/Dólar Historico.xlsx"
    
    print("")
    print("CARPETA SELECCIONADA:", nombre_carpeta)
    print("Abriendo archivos 'xlsx.'...")
    print("")
    print("Abriendo Presupuesto...")
    print("")
    
    # PRESUPUESTO
     # Con OpenPyXl
    presupuesto_openpyxl = openpyxl.load_workbook(ruta_presupuesto, data_only=True)
    hoja_resumen_pesos = presupuesto_openpyxl["RESUMEN DE OBRA PESOS"]
    hoja_resumen_dolares = presupuesto_openpyxl["RESUMEN DE OBRA U$D"]
    
     # Con WingsXl
    presupuesto_xw = xw.Book(ruta_presupuesto)
    resumen_pesos = presupuesto_xw.sheets["RESUMEN DE OBRA PESOS"]
    resumen_dolares = presupuesto_xw.sheets["RESUMEN DE OBRA U$D"]
    print("Abriendo Gastos...")
    print("")

    # GASTOS
     # Con openPyXl
    gastos_open = openpyxl.load_workbook(ruta_gastos)
    
     # Con Pandas
    gastos = pd.read_excel(ruta_gastos, sheet_name="GASTOS DE OBRA", skiprows= 1)
    
else:
    print("No se seleccionó ninguna carpeta.")
    
    


# In[ ]:


# 3.bis - REINICIO

# # # guardo archivos excel y leo nuevamente si falla. 
# # Intento una vez. Si continúa el error. Aviso y cierro el programa

def reiniciar_archivos_excel(presupuesto_openpyxl, gastos_open):
    
    print("REINICIANDO ARCHIVOS...")
    
    # Ruta de los archivos a reiniciar
    ruta_presupuesto = f"{ruta}/Antecedente/PRESUPUESTO GENERAL - {nombre_carpeta}.xlsx"
    ruta_gastos = f"{ruta}/Antecedente/GASTOS - {nombre_carpeta}.xlsx"

    # Reescribir PRESUPUESTO
    
    presupuesto_openpyxl.save(ruta_presupuesto)  # Guardar los cambios y sobrescribir el archivo existente
    presupuesto_openpyxl = openpyxl.load_workbook(ruta_presupuesto)  # Guardar los cambios y sobrescribir el archivo existente
    
    # Reescribir GASTOS
    gastos_open.save(ruta_gastos)  # Guardar los cambios y sobrescribir el archivo existente
    gastos_open = openpyxl.load_workbook(ruta_gastos)
    
    print("DONE")

    return presupuesto_openpyxl, gastos_open


# In[ ]:


# 5 - DATAFRAMES VARIABLES

# - LEER EXCEL DÓLAR -

#FUNCIÓN: 
# Crea un Data Frame por cada tipo de dólar
# Se almacena en un diccionario

print(" # CREAR DATAFRAME DÓLAR... ")
print("")

def DOLAR(): 
    variables_dolar = ["oficial", "mep", "blue"] 
    dolar = {}    
    
    for var_dolar in variables_dolar:
        
        df = pd.read_excel(ruta_variables, sheet_name="DÓLAR HISTÓRICO", skiprows = 1,
                          usecols = ["fecha", f"{var_dolar}-promedio"])
        
        
        df = df.rename(columns={f"{var_dolar}-promedio": var_dolar})
        
        dolar[var_dolar.upper()] = df 

                           
    return dolar
    

dolar = DOLAR() #call the function

    
# - LEER EXCEL CAC -

#FUNCIÓN: 
# Crea un Data Frame
# Almaceno en dictionary

print(" # CREAR DATAFRAME CAC... ")
print("")

def CAC(): 
    variables_cac = ["mano de obra", "costo construcción", "materiales"]
    cac = {}
    
    for variable in variables_cac:
        
        df = pd.read_excel(ruta_variables, sheet_name="CAC",
                          skiprows=1, usecols = ["fecha", f"indice-{variable}"])
        
        nombre = variable.replace(" ", "_").upper()
        
        df = df.rename(columns={f"indice-{variable}": nombre})
        cac[nombre] = df
    return cac
                
cac = CAC() #Call the function


# In[ ]:


# 8 - ESTRUCTURA DE BASE DE DATOS

# 

# Contenedores:
# CONTENEDOR [N°RUBRO] [CATEGORIA] [PRESUPUESTO/PAGO] [CLASE]

#1. Almacena N° de Rubro
CONTENEDOR = {} 

#2. Nro de rubro - El primer item siempre será "PRESUPUESTO"
RUBROS = [] 
NAME_RUBROS = [] #Nombre de rubro. A modo de referencia

#3. Categorías. 
MANO_DE_OBRA = "MANO_DE_OBRA"
MATERIALES = "MATERIALES"
VARIOS = "VARIOS"  
ADICIONAL = "ADICIONAL"

#Alojo las variables en lista. Para facilitar busqueda.
CATEGORIAS = [VARIOS, MANO_DE_OBRA, MATERIALES, ADICIONAL]

#4. Diferencio entre:
# Presupuesto
# Gastos

#5. Clase donde alojo las variables importantes
class Presupuesto:
    def __init__(self, nombre_presupuesto, rubro, subrubro, fecha_inicial, 
                 indice, indice_base, indice_cac_base, dolar_inicial, 
                 presupuesto_original, presupuesto_iva, presupuesto_dolares, presupuesto_mep, presupuesto_blue,
                 estado, presupuesto_actualizado, base_presupuesto_actualizado, presupuesto_dolares_actualizado,
                 saldo, saldo_dolares):
           
        self.nombre_presupuesto = nombre_presupuesto     #1
        self.rubro = rubro                               #2
        self.subrubro = subrubro                         #3
        self.fecha_inicial = fecha_inicial               #4
        self.indice = indice                             #5
        self.indice_base = float(indice_base)            #6
        self.indice_cac_base = indice_cac_base           #7  ## Calculado
        self.dolar_inicial = dolar_inicial               #8  ## Calculado
        self.presupuesto_original = presupuesto_original #9 
        self.presupuesto_iva = presupuesto_iva           #10 
        self.presupuesto_dolares = presupuesto_dolares   #11 ## Calculado
        self.presupuesto_mep = presupuesto_mep           #12 
        self.presupuesto_blue = presupuesto_blue         #13 
        self.estado = estado                             #14 
        self.presupuesto_actualizado = presupuesto_actualizado  # 15
        self.base_presupuesto_actualizado = base_presupuesto_actualizado  #16
        self.presupuesto_dolares_actualizado = presupuesto_dolares_actualizado  #17
        self.saldo = saldo
        self.saldo_dolares = saldo_dolares

class Pago: 
    def __init__(self, posicion, posicion_opy, fecha, fecha_cac, subrubro, base_cac, oficial, mep, blue, monto_pesos,monto_pesos_act, monto_dolares, monto_dolares_act):
        self.posicion = posicion
        self.posicion_opy = posicion_opy
        self.fecha = fecha
        self.fecha_cac = fecha_cac
        self.subrubro = subrubro
        self.base_cac = base_cac
        self.oficial = oficial
        self.mep = mep
        self.blue = blue
        self.monto_pesos = monto_pesos
        self.monto_pesos_act = monto_pesos_act
        self.monto_dolares = monto_dolares
        self.monto_dolares_act = monto_dolares_act
        
class Saldo:
    def __init__(self, fecha, subrubro, saldo_pesos, saldo_dolares):
        self.fecha = fecha
        self.subrubro = subrubro
        self.saldo_pesos = saldo_pesos
        self.saldo_dolares = saldo_dolares
        


# In[ ]:


# 9 - DIFERENCIO HOJA CORRESPONDIENTE A RUBRO 

# 
# GENERO ESTRUCTURA INTERNA EN CADA UNO DE ELLOS.

# Las Hojas Contenidas entre "PRESUPUESTO" y "FIN DEL PRESUPUESTO"
# La posición de cada "hoja", dentro del array, actuará como clave  
# CONTENEDOR[n° = posición de cada hoja]


def Insertar_Rubros_A_Contenedor():
    

    #lee hojas de archivo
    hojas = presupuesto_openpyxl.sheetnames 
    nombre_hojas = [] 
    
  
    adentro = False
    
    # Solo almacena las que están entre RESUMEN DE OBRA Y FIN PRESUPUESTO
    for hoja in hojas: 
        if hoja == "RESUMEN DE OBRA PESOS":
            adentro = True
            
        elif hoja == "FIN PRESUPUESTO": 
            adentro = False
        
        if adentro:
            NAME_RUBROS.append(hoja)
            
    # Almaceno en database según posición en array
    for nro, i in enumerate(NAME_RUBROS):        
        if i == "RESUMEN DE OBRA PESOS":
            RUBROS.append(nro)
            CONTENEDOR[nro] = "RESUMEN DE OBRA PESOS"
            print("RESUMEN DE OBRA:")            
            
        else:
            
            # RESUMEN DE OBRA 
            presupuesto = []
            
            # PAGOS
            pagos = []
            
            # SALDO
            saldo = []
                
            # POR CADA CATEGORIA ASIGNO UN DIC VACIO EN PRESUPUESTO Y PAGOS
            CONTENEDOR[nro] = {
                "MANO_DE_OBRA": {
                    "saldo": saldo.copy(),
                    "presupuesto": presupuesto.copy(),
                    "pagos": pagos.copy()
                },
                "MATERIALES": {
                    "saldo": saldo.copy(),
                    "presupuesto": presupuesto.copy(),
                    "pagos": pagos.copy()
                },
                "VARIOS": {
                    "saldo": saldo.copy(),
                    "presupuesto": presupuesto.copy(),
                    "pagos": pagos.copy()
                },
                "ADICIONAL": {
                    "saldo": saldo.copy(),
                    "presupuesto": presupuesto.copy(),
                    "pagos": pagos.copy()
                }
            }
            
    return 
        
primer_paso = Insertar_Rubros_A_Contenedor()
print("  Estructura Interna Generada")
print("")


# In[ ]:


# 10 - EXTRAIGO CUANTOS PRESUPUESTOS TENGO POR RUBRO. 
def cantidad_rubros(presupuesto_openpyxl, name):
    
    # METODO SIMPLE, busco por posicion
    hoja = presupuesto_openpyxl[name]
    dato = hoja["L2"].value
    
    # Si es nan o None
    if dato == None or dato == "nan":
        
        # METODO TABLA. Busco por nombre tabla. 
        
        tablas = []
        tabla = str

        for tabla in hoja.tables.values():
            tablas.append(tabla.displayName) 

        for nombre_tablas in tablas:
            patron = re.compile(r'cantidad', re.IGNORECASE)
            coincidencias = patron.findall(nombre_tablas)
            if coincidencias:
                coincidencias = str(coincidencias)
                tabla = nombre_tablas
                print("tabla", tabla)
                print("")        

                resumen = hoja.tables[tabla]
                # Obtener el rango de la columna específica (por ejemplo, 'Columna1')
                dato = resumen.ref[1]

            dato = hoja["L2"].value
            
    return dato

    

# 
 

def Rows_Presupuestos(reinicios, presupuesto_openpyxl, gastos_open):
    
    #Array donde alojo el dato
    cantidad_presupuestos = []
    
    
    #Por cada rubro alojado previamente
    #Registro Dato 'Cuantos presupuesto' de la hoja de rubro
    for name in NAME_RUBROS: 

        if name == "RESUMEN DE OBRA PESOS":
            cantidad_presupuestos.append(0) #Spot 0. Debe tener mismo size que List_Rubros
            print("")
            print("Cantidad de presupuestos por rubro: ")
            

        else:
            dato = cantidad_rubros(presupuesto_openpyxl, name)
            
            if dato == None or dato == "nan":
            
                if reinicios == 0:   
                    # SI SIGUE SIENDO NAN, reinicio el programa. 
                    print("")
                    print("Problemas con excel")
                    print("REINICIANDO EL PROGRAMA")
                    reiniciar = reiniciar_archivos_excel(presupuesto_openpyxl, gastos_open)
                    presupuesto_openpyxl, gastos_open = reiniciar
                    reinicios +=1
                    again = Rows_Presupuestos(reinicios, presupuesto_openpyxl, gastos_open)
                    print("")

                else:
                    ventana = tk.Tk()
                    ventana.withdraw()  # Ocultar la ventana principal

                    # Mostrar el mensaje de error en una ventana emergente
                    messagebox.showerror("Error", "NO SE LOGRÓ LEER ARCHIVO EXCEL\nGUARDAR EXCEL E INTENTAR DE NUEVO")

                    # Mostrar un mensaje adicional si lo deseas
                    # messagebox.showinfo("Información", "Haz clic en Aceptar para cerrar el programa")

                    ventana.destroy()  # Cerrar la ventana emergente

                    print("")
                    print("NO SE LOGRÓ LEER ARCHIVO EXCEL")
                    print("GUARDAR EXCEL PRESUPUESTO E INTENTAR DE NUEVO")
                    sys.exit()
                    
        
            cantidad_presupuestos.append(dato)  
            
            print(name, ": ", dato)
    return cantidad_presupuestos

reinicios = 0

cantidad_presupuestos = Rows_Presupuestos(reinicios, presupuesto_openpyxl, gastos_open)


# In[ ]:


# 11 - ACCEDER A CADA PRESUPUESTO Y EXTRAER DATOS

def Crear_Presupuestos():
    
    print("LECTURA DE PRESUPUESTOS")
    
    for i, rubro in enumerate(NAME_RUBROS):
        
        if rubro == "RESUMEN DE OBRA PESOS": 
            print("Extracción de datos de 'Resumen' en cada rubro...")
            print("")

        else:
            print("-----------------------------------------------")
            print("")
            print("RUBRO", rubro.upper())
            print("")
            
            presupuestos = cantidad_presupuestos[i]  
            #si no hay presupuesto
            if  presupuestos == 0:
                print(rubro, "Empty")
                print("")
            
            #si hay presupuesto
            else:
                
                #leo pagina Excel. 
                #EXCEL PAGE == HOJA RUBRO
                excel_page = pd.read_excel(ruta_presupuesto,
                                           sheet_name = rubro,
                                           usecols = ("CATEGORÍA","RUBRO", "SUB_RUBRO", "PRESUPUESTO", "FECHA", "INDICE", "BASE", "MONTO", "MONTO + IVA", "ESTADO"),
                                           nrows = cantidad_presupuestos[i])
                
                # ----------------------------------------------------------- # 
                # Esta parte se podría obviar con la restricción desde excel ...

                # "ARREGLO" DATATYPE EN BASE. DEBERÍA SER FLOAT, no str
                if excel_page['BASE'].dtype == 'object':
                    
                    # Limpia todo catacter que no sea número
                    excel_page['BASE'] = excel_page['BASE'].str.replace('[^\d.]', '', regex=True)

                    # Reemplaza ',' por '.' en la columna "BASE" y convierte a float
                    excel_page['BASE'] = excel_page['BASE'].str.replace(',', '.', regex=True).astype(float)
        
            
            
                #CAMBIO FORMATO DE COLUMNA FECHA
                excel_page['FECHA'] = pd.to_datetime(excel_page['FECHA']).dt.strftime('%Y-%m-%d')
                
                #Me facilito el index del dataframe
                index = excel_page.index 
                                                                # Hasta aquí 
                # ----------------------------------------------------------- # 
                
                
                # EXTRAIGO LOS DATOS A INGRESAR A MI DATABASE
                
                hoja = presupuesto_xw.sheets(rubro)
                for linea in range(presupuestos):
                    
                    print(f"PRESUPUESTO N°{(linea + 1)}")
                    
                    #Fila
                    fila = excel_page.loc[excel_page.index == linea]
                    nro_fila = linea + 2

                    # NOMBRE
                    nombre_presupuesto = str((excel_page.loc[index == linea, "PRESUPUESTO"]).values)
                    
                    # FECHA
                    fecha_inicial = excel_page.loc[index == linea, "FECHA"].iloc[0]
                    try:
                        fecha_inicial = datetime.strptime(fecha_inicial, "%Y-%m-%d").date()
                        
                    except TypeError or ValueError:
                        print(" Formato fecha incorrecto!")
                        return
                    
                    fecha_cac = fecha_inicial.replace(day=1) - relativedelta(months=1) # FECHA CAC MES ANTERIOR
                    
                    print(f" fecha {fecha_inicial} y nombre {nombre_presupuesto}")
                    
                    # CATEGORIA 
                    categoria = fila.at[fila.index[0], "CATEGORÍA"]
                    categoria = categoria.replace(" ", "_")
                    
                    # Si en el nombre se menciona la palabra adicional, 
                    # la categoría será ADICIONAL
                    patron = re.compile(r'adicional', re.IGNORECASE)
                    coincidencias = patron.findall(nombre_presupuesto)
                    if coincidencias:
                        coincidencias = str(coincidencias[0]).upper()
                        print("El presupuesto se tomara como 'Adicional'")

                        categoria = coincidencias
                        print(" Categoria", categoria)
                        print("")        
                    
                    print(" Categoría :", categoria)
             
                        
                    # RUBRO
                    rubro = fila.at[fila.index[0], "RUBRO"]
                    
                    # SUB-RUBRO
                    subrubro = fila.at[fila.index[0], "SUB_RUBRO"]
                    
                    print(" RUBRO", rubro, "SUB-RUBRO:", subrubro)
                    
                    
                    # INDICE
                    indice = str(excel_page.loc[index == linea, "INDICE"].values[0]).replace(" ", "_")
                    
                    # BASE 
                    indice_base = excel_page.loc[index == linea, "BASE"].values[0]
                    
                    # MONTO
                    presupuesto_original = excel_page.loc[index == linea, "MONTO"].values
                    
                    # MONTO + IVA
                    presupuesto_iva = excel_page.loc[index == linea, "MONTO + IVA"].values
                    
                    # ESTADO 
                    estado = excel_page.loc[index == linea, "ESTADO"].values
                    
                    # Extraidos de "Tabla Resumen"
                    # ---------------------------------------------------------------------
                    
       
                    #CAC BASE
                    indice_cac_base = float
                    dolar_inicial = 0
                    
                    
                    # DOLAR INICIAL - Según indice si es dolar. Blue, para indice cac
                    if indice == "OFICIAL" or indice == "BLUE" or indice == "MEP":
                        
                        if (np.isnan(indice_base)) == False:
                            dolar_inicial = indice_base

                        # Si la Base Está "VACIA", completo segun indice
                        else:
                            dolar_inicial = dolar[indice].loc[dolar[indice]["fecha"] == (pd.to_datetime(fecha_inicial)), f"{indice.lower()}"].values[0]
                            indice_base = dolar_inicial
                            print(" INDICE NAN: DOLAR", dolar_inicial)
                            #Guardo valor en presupuesto con xlwings
                            hoja.range(f"G{nro_fila}").value = dolar_inicial
                            
                        # Si el indice es dolar:
                        # Aún necesito el indice cac:
                        indice_cac_base = cac["COSTO_CONSTRUCCIÓN"].loc[cac["COSTO_CONSTRUCCIÓN"]["fecha"] == (pd.to_datetime(fecha_cac)), "COSTO_CONSTRUCCIÓN"]
                        indice_cac_base = indice_cac_base.values[0]
                        
                        # SI el índice no es "dolar", de todas maneras necesito ese valor. 
                        if not indice == "BLUE":
                            dolar_blue = dolar["BLUE"].loc[dolar["BLUE"]["fecha"] == (pd.to_datetime(fecha_inicial)), "blue"].values[0]
                            presupuesto_en_blue = presupuesto_original/dolar_blue
                    
                    # SI EL INDICE ES CAC
                    else: 
                        # El dolar inicial es Blue
                        dolar_inicial = dolar["BLUE"].loc[dolar["BLUE"]["fecha"] == (pd.to_datetime(fecha_inicial)), "blue"].values[0]
                        
                        # Y CaC Inicial es indice Base
                        if (np.isnan(indice_base)) == False:
                            indice_cac_base = indice_base
                        
                        # Si Indice Base está vacío: busco el indicado
                        # Guardo valro en presupuesto con xlwings
                        else:
                            
                            indice_cac_base = cac[indice].loc[cac[indice]["fecha"] == (pd.to_datetime(fecha_cac)), indice].values[0]
                            hoja.range(f"G{nro_fila}").value = indice_cac_base
                            indice_base = indice_cac_base
                            
                            print(" INDICE BASE NAN: CAC", indice_cac_base)
                          
                    print(" Indice: ", indice, "Base:", indice_base)
                    print(" Dolar inicial:", dolar_inicial)
                    
                    # PRESUPUESTO EN DÓLARES
                    presupuesto_dolares = int(presupuesto_original/ dolar_inicial)
                    
                    # Presupuesto en MEP
                    mep = dolar["MEP"].loc[dolar["MEP"]["fecha"] == (pd.to_datetime(fecha_inicial)), "mep"]

                    presupuesto_mep = presupuesto_original / float(mep)
                    
                    # Presupuesto Blue
                    
                    blu = dolar["BLUE"].loc[dolar["BLUE"]["fecha"] == (pd.to_datetime(fecha_inicial)), "blue"]
                    presupuesto_blue = presupuesto_original / float(blu)
                    
                    # CAC
                    indice_cac_base # El gasto también tiene un índice
                    base_presupuesto_actualizado = indice_base
                            
                    # ACTUALIZACIÓN
                    presupuesto_actualizado = presupuesto_original
                    presupuesto_dolares_actualizado = presupuesto_dolares
                    
                    # SALDOS
                    saldo = presupuesto_original
                    saldo_dolares = presupuesto_dolares
                    
                    
                    # GENERO CLASE

                    clave = nombre_presupuesto
                    clave = Presupuesto(nombre_presupuesto, rubro, subrubro, fecha_inicial,
                                        indice, indice_base, indice_cac_base, dolar_inicial, 
                                        presupuesto_original, presupuesto_iva, presupuesto_dolares, presupuesto_mep, presupuesto_blue,
                                        estado, presupuesto_actualizado, base_presupuesto_actualizado, presupuesto_dolares_actualizado,
                                        saldo, saldo_dolares)

                    
                    # En última instancia, lo alojo en database según categoría 
                    # (lista categoria definida junto con estructura de database)
                    CONTENEDOR[i][categoria]["presupuesto"].append(clave)

                    print(" Indice CAC:", indice_cac_base)
                    print(" Pto. Original", presupuesto_original)
                    print(" Pto. Dólares", presupuesto_dolares)
                    print("")

    return

try:
    presupuestos = Crear_Presupuestos()
except Exception as e:
    
    print(f"El programa fallá en {e}")
    traceback.print_exc()  
    print("")
    input("Presiona 'Enter' para cerrar terminal")
    sys.exit()


# In[ ]:


# 12 - BACKUP

# ANTES DE MODIFICAR ARCHIVOS - x
# SE BUSCAN ARCHIVOS .XLSX EN CARPETA,
# SE DUPLICAN Y SE ARCHIVAN EN ANTECEDENTES

if ruta:
    ruta = ruta.replace('\\', "/")
    # Listar los archivos Excel en la carpeta seleccionada
    archivos_excel = [archivo for archivo in os.listdir(ruta) if archivo.endswith(".xlsx")]

    # Obtener la fecha actual y formatearla como cadena (por ejemplo, "2023-10-04")
    fecha_actual = datetime.now().strftime("%Y-%m-%d")
    
    # Duplicar y guardar los archivos en la carpeta "Antecedente" con la fecha en el nombre
    print("")
    print("COPIA A ANTECEDENTE: ")
    for archivo in archivos_excel:
        if not archivo.startswith("~$"):
            origen = f"{ruta}\\{archivo}"
            nombre_sin_extension, extension = os.path.splitext(archivo)
            nuevo_nombre = f"{nombre_sin_extension}_{fecha_actual}{extension}"
            destino = f"{ruta}\Antecedente\\{nuevo_nombre}"

            shutil.copy2(origen, destino)
            
            print(nuevo_nombre)
    print("")
else:
    print("No se seleccionó ninguna carpeta.")
      



# In[ ]:


# 13 - PROCESAR GASTOS. Almaceno cada gasto dentro de la estructura

def Actualizacion_general(gastos_open):
    
    gastos_xw = xw.Book(ruta_gastos)
    hoja_gastos = gastos_xw.sheets("GASTOS DE OBRA")
    
    # Para trabajar con openpyxl
    col_oficial = "N"
    col_mep = "O"
    col_blue = "P"
    
    col_saldo_pesos = "Q"
    col_saldo_blue = "R"
    col_estado = "S"
    
    index_menos_ultima_fila = len(gastos.index) - 1
    print("CANTIDAD DE SALIDAS EN GASTOS: ", index_menos_ultima_fila)
    print("")
    
    for index in range(index_menos_ultima_fila):
        print("")
        print(" ----- Gasto #", index)
        
        index_open = index + 3 #Index para aplicar cambios con openpyxl
        
        
        # ------------------------- CHEQUEAR QUE COINCIDAN! ---------------------- #
        #VARIABLES NECESARIAS EN TODAS LAS FUNCIONES:
        print("")
        
        moneda = str(gastos.loc[gastos.index[index], "MONEDA"]).upper().replace(" ", "")
        
        medio = str(gastos.loc[gastos.index[index], "MEDIO DE PAGO"]).upper().replace(" ", "")
        
        print(f" Moneda: {moneda} / Medio: {medio}")
        
 
        fecha = gastos.loc[gastos.index[index], "FECHA"]

        fecha_cac = fecha.replace(day=1) - relativedelta(months=1) # FECHA CAC MES ANTERIOR
        #print("fecha_cac", fecha_cac)
        
        print(" Fecha", (fecha.date()))
        
        categoria = str(gastos.loc[gastos.index[index], "CATEGORÍA"]).replace(" ", "_").upper()
        
        print(" Categoria", categoria)
 
        rubro = gastos.loc[gastos.index[index], "RUBRO"]
        if isinstance(rubro, float) and np.isnan(rubro):
            rubro = "-"# Código si 'rubro' es NaN
            
        subrubro = str(gastos.loc[gastos.index[index], "SUB-RUBRO"])
        if subrubro == "nan":
            subrubro = None
            
        print(f" Rubro: {rubro} / Sub: {subrubro}")
        
        descripcion = str(gastos.loc[gastos.index[index], "DESCRIPCIÓN"])
        
        print(" Descripcion", descripcion)
        
        ## DOLAR 
        oficial = dolar["OFICIAL"].loc[dolar["OFICIAL"]["fecha"] == fecha, "oficial"].values
        hoja_gastos.range(f"{col_oficial}{index_open}").value = int(oficial)
        
        mep = dolar["MEP"].loc[dolar["MEP"]["fecha"] == fecha, "mep"].values
        hoja_gastos.range(f"{col_mep}{index_open}").value = int(mep)
        
        blue = dolar["BLUE"].loc[dolar["BLUE"]["fecha"] == fecha, "blue"].values
        hoja_gastos.range(f"{col_blue}{index_open}").value = int(blue)
        
        
        ## CAC  
        # Si en el nombre se menciona la palabra adicional, 
        # la categoría será ADICIONAL
        patron = re.compile(r'adicional', re.IGNORECASE)
        coincidencias = patron.findall(descripcion)
        if coincidencias:
            coincidencias = str(coincidencias[0]).upper()
            print("El presupuesto se tomara como 'Adicional'")

            categoria = coincidencias
            print("Categoria", categoria)
            print("")     
            
             
        # MONTO
        monto = gastos.loc[gastos.index[index], "MONTO"]
        monto_dolares = 0     
                
        # SI TIENE ASIGNADO UN RUBRO BUSCO EL NUMERO:
        nro_rubro = None
        if not rubro == "-":
            
            #LIMPIO TODO CARACTER QUE NO SEA UN NÚMERO DEL NOMBRE DEL RUBRO
            def limpiar_nro_de_rubro(nro_rubro):
                numeros = ''.join(caracter for caracter in nro_rubro if caracter.isdigit())
                return int(numeros)
        
            try:
                nro_rubro = limpiar_nro_de_rubro(rubro)
                
            except ValueError:
                
                print("VALUE ERROR")
                # Manejar la excepción, por ejemplo, asignar un valor predeterminado a nro_rubro
                nro_rubro = "-"  # O cualquier otro valor predeterminado
                
                continue
        

        if not nro_rubro == None:
            
            ## ALOJO GASTO EN ESTRUCTURA 
            
            posicion = index
            posicion_opy = index_open
                        
            monto_pesos = 0
            monto_dolares = 0
            
            presupuestos = CONTENEDOR[nro_rubro][categoria]["presupuesto"]
            
            indice = 0
            #print("_______________")
            
            # SI HAY PRESUPUESTOS ALMACENADOS EN EL RUBRO            
            if not len(presupuestos) == 0:
                
                for presupuesto in presupuestos:
                    sub = (presupuesto.subrubro)
                    #print("SUB, ", sub)
                    
                    if sub == subrubro:
                        #print(f"Coinciden los sub? {sub};{subrubro}")
                        indice = presupuesto.indice
                        #print("INDICE;", indice)
                        break

                    else:
                        indice = presupuesto.indice
                        #print("INDICE;", indice)
                        
            if not len(presupuestos) == 0:
                indice_presu = presupuestos[0].indice
                if indice_presu == "MANO_DE_OBRA" or indice_presu == "MATERIALES":
                    try:
                        base_cac = float(cac[indice_presu].loc[cac[indice_presu]["fecha"] == fecha_cac, indice_presu].values[0])   
                    except pd.errors.OutOfBoundsDatetime:
                        fecha_cac = fecha.replace(day=1) - relativedelta(months=2) # FECHA CAC segundo MES ANTERIOR
                        base_cac = float(cac[indice_presu].loc[cac[indice_presu]["fecha"] == fecha_cac, indice_presu].values[0])   

                else:
                    try:
                        base_cac = float(cac["COSTO_CONSTRUCCIÓN"].loc[cac["COSTO_CONSTRUCCIÓN"]["fecha"] == fecha_cac, "COSTO_CONSTRUCCIÓN"].values[0])   
                    except pd.errors.OutOfBoundsDatetime:
                        fecha_cac = fecha.replace(day=1) - relativedelta(months=2) # FECHA CAC segundo MES ANTERIOR
                        base_cac = float(cac["COSTO_CONSTRUCCIÓN"].loc[cac["COSTO_CONSTRUCCIÓN"]["fecha"] == fecha_cac, "COSTO_CONSTRUCCIÓN"].values[0])   

            # ME ASEGURO DE TENER EL MONTO CORRECTO
            if moneda == "DÓLAR" or moneda == "DÓLARES" or moneda == "DOLAR":
                
                # INDICE ES OFICIAL
                if indice == "OFICIAL":
                    monto_dolares = monto
                    monto = monto_dolares * oficial
                    #print("OFICIAL")
                    
                elif indice == "MEP":
                    monto_dolares = monto
                    monto = monto_dolares * mep
                    #print("MEP")
                    
                else:
                    monto_dolares = monto
                    # Pagos por banco son MEP / En efe son Blue
                    if medio == "BANCO":
                        monto = monto * mep
                        #print("BANCO-MEP")
                    else:
                        monto = monto * blue
                        #print("BANCO-BLUE")
                        
            # Si la moneda es PESOS
            else:
                # INDICE ES OFICIAL
                if indice == "OFICIAL":
                    monto_dolares = monto / oficial
                elif indice == "MEP":
                    monto_dolares = monto / mep
                else:
                    monto_dolares = monto
                    # Pagos por banco son MEP / En efe son Blue
                    if medio == "BANCO":
                        monto_dolares = monto / mep
                    else:
                        monto_dolares = monto / blue
                        
            print(f" Indice: {indice} / Base {base_cac}")
            
            monto_pesos = monto
            monto_pesos_act = monto_pesos
            monto_dolares_act = monto_dolares
                    
                    
            # GENERO CLASE
            clave = fecha # El presupuesto se aloja con el nombre como clave
            clave = Pago(posicion, posicion_opy, fecha, fecha_cac, subrubro, base_cac, oficial, mep, blue, monto_pesos, monto_pesos_act, monto_dolares, monto_dolares_act)


            # En última instancia, lo alojo en database según categoría 
            # (lista categoria definida junto con estructura de database)
            try:
                CONTENEDOR[nro_rubro][categoria]["pagos"].append(clave)
            except KeyError:
                print("ERROR EN CLASIFICACION GASTO!")
                         
            print("")
            print(" Monto: $", monto)
            print(" En Dolares: U$", int(monto_dolares))
            print("")
            print("------------------------------")
            
            
            
    gastos_xw.save()


try:
    prueba = Actualizacion_general(gastos_open)
    
except Exception as e:
    
    print(f"El programa fallá en {e}")
    traceback.print_exc()  
    print("")
    input("Presiona 'Enter' para cerrar terminal")
    sys.exit()



# In[ ]:


# 14.2.b - CASO N°3 EN PESOS
# Hay presupuestos y pagos.

def Caso_3_pesos(hoja_gastos, presupuestos_disponibles, pagos_realizados):
    
    col_pesos = "Q"
    col_dolares = "R"
    print("")
    print("ACTUALIZA EN PESOS")
    print("")


    resto = 0   # afectará si el pago es mayor al presupuesto/saldo
    cac_resto = 0
    
    # MIENTRAS TENGA PAGOS
    for posicion_pago, pago in enumerate(pagos_realizados):
        
        #Si el pago tiene subrubro, uso solo los presupuestos con el mismo subrubro
        sub_rubro = (pago.subrubro)
        
        
        if not (sub_rubro == "nan") or not (sub_rubro == None) or not(sub_rubro == np.nan):
        
            presupuestos_con_subrubro = []
            for presu in presupuestos_disponibles:
                
                subr_presu = presu.subrubro
                if not (subr_presu == "nan") or not(subr_presu == None) or not(subr_presu == np.nan) and sub_rubro == subr_presu:
                    presupuestos_con_subrubro.append(presu)
                    
            presupuestos_disponibles = presupuestos_con_subrubro
                    
        # Si el pago ya fue analizado:
        if pago.monto_pesos_act <= 0:
            #print("continuamos")
            continue
            
        else:
            # llevo cuenta de los presupuestos usados
            presupuestos_agotados = sum(1 for presupuesto in presupuestos_disponibles if presupuesto.presupuesto_actualizado <= 0)


            for posicion_presu, presupuesto in enumerate(presupuestos_disponibles):

                # Si agoté mis presupuestos, acumulo los pagos como saldo.
                if presupuestos_agotados == len(presupuestos_disponibles):

                    #print("Presupuestos agotados")
                    #print("") #CASO N°1

                    #ACUMULO PAGO, EN PESOS Y EN DOLARES, COMO SALDO NEGATIVO
                    acumulado_pesos = 0
                    while posicion_pago < len(pagos_realizados):

                        pago = pagos_realizados[posicion_pago]
                        monto_pesos = pago.monto_pesos_act

                        if not resto == 0:
                            monto_pesos = monto_pesos + resto
                            resto = 0

                        posicion_opy = pago.posicion_opy
                        acumulado_pesos = acumulado_pesos + monto_pesos
                        #print("acumulado_pesos", acumulado_pesos)

                        hoja_gastos.range(f"{col_pesos}{posicion_opy}").value = int(-(acumulado_pesos))
                        posicion_pago +=1
                    return


                # Si el presupuesto ya fue analizado
                if presupuesto.presupuesto_actualizado <= 0:
                    continue

                else:

                    #print("Pto. N°", posicion_presu)
                    print(presupuesto.presupuesto_actualizado)
                    #print("")


                    #print("Pago", posicion_pago, pago.monto_pesos_act)

                    fecha_pago = pago.fecha
                    monto_pesos = pago.monto_pesos_act
                    indice_actual = pago.base_cac

                    # CHEQUEO CAC

                    if presupuesto.indice == "MANO_DE_OBRA" or presupuesto.indice == "MATERIALES":
                        try:
                            fecha_cac = fecha_pago.replace(day=1) - relativedelta(months=1) # FECHA CAC segundo MES ANTERIOR
                            base_cac = float(cac[presupuesto.indice].loc[cac[presupuesto.indice]["fecha"] == fecha_cac, presupuesto.indice].values[0])   

                        except pd.errors.OutOfBoundsDatetime:
                            fecha_cac = fecha_pago.replace(day=1) - relativedelta(months=2) # FECHA CAC segundo MES ANTERIOR
                            base_cac = float(cac[presupuesto.indice].loc[cac[presupuesto.indice]["fecha"] == fecha_cac, presupuesto.indice].values[0])   
                            
                    #    print(f"INDICE: MANO DE OBRA | MATERIALES  {base_cac}")
                    
                    # SI EL PAGO ES ANTICIPO
                    elif indice_actual == 0:
                        indice_actual = presupuesto.indice_cac_base
                            
                    else:
                        fecha_cac = fecha_pago.replace(day=1) - relativedelta(months=2) # FECHA CAC segundo MES ANTERIOR
                        base_cac = float(cac["COSTO_CONSTRUCCIÓN"].loc[cac["COSTO_CONSTRUCCIÓN"]["fecha"] == fecha_cac, "COSTO_CONSTRUCCIÓN"].values[0])   
                        
                    posicion_opy = pago.posicion_opy

                    if not resto == 0:

                       # print("HAY RESTO", resto)
                      #  print("")
                        if cac_resto > (presupuesto.indice_cac_base):

                            #ACTUALIZAR EL PRESUPUESTO

                            # ACTUALIZO Y RESTO EL RESTO ~cuak~
                            presupuesto_pesos = presupuesto.presupuesto_actualizado
                            indice_cac_base = presupuesto.indice_cac_base

                            # Actualizar

                            indice_cac = cac_resto / indice_cac_base
                            actualizacion = presupuesto_pesos * indice_cac

                            print(f"{presupuesto_pesos} * ({cac_resto} / {indice_cac_base})")
                            print("=", actualizacion)

                            presupuesto_pesos = presupuesto.presupuesto_actualizado
                            print("presupuesto_pesos", presupuesto_pesos)
                            actualizacion = presupuesto_pesos - resto
                            print(f"actualizacion = {presupuesto_pesos} - {resto}")
                            print("=", actualizacion)

                        else:
                            presupuesto_pesos = presupuesto.presupuesto_actualizado
                            print("presupuesto_pesos", presupuesto_pesos)
                            actualizacion = presupuesto_pesos - resto
                            print(f"actualizacion = {presupuesto_pesos} - {resto}")
                            print("=", actualizacion)


                        if actualizacion <= 0:

                            resto = -(actualizacion)
                            cac_resto = indice_actual 
                            presupuesto.presupuesto_actualizado = 0

                            # Para trasladar a Excel debería sumar el saldo entero.
                            # Debo sumar los presupuestos restantes. 
                            # Y restarles el saldo.
                            suma = 0
                            for presu in presupuestos_disponibles:

                                monto = presu.presupuesto_actualizado
                                cac__ = presu.indice_cac_base
                                
                                aux__ = indice_actual / cac__
                                monto = monto*aux__
                                
                                #Actualizo Los presupuesos.
                                suma = suma + monto
                                
                            suma = suma - resto
                            
                            hoja_gastos.range(f"Q{posicion_opy}").value = int(suma)
                      #      print("SE SUMAN Ptos RESTANTES SI PAGO>PTO")
                      #      print(f"N°{p} SUMA DE PRESUPUESTOS:", suma)
                            break

                        else:
                            presupuesto.presupuesto_actualizado = actualizacion
                            resto = 0


                    # ACTUALIZO Y RESTO PAGO 
                    presupuesto_pesos = presupuesto.presupuesto_actualizado
                    indice_cac_base = presupuesto.indice_cac_base

                    # Actualizar

                    ## Evito actualizar si el pago es anterior al cac base 
                    ## numpy.dtypes.Float64DType
                    if indice_actual == None:
                        indice_actual = indice_cac_base
                        print("ANTICIPO")
                        
                    if indice_actual < indice_cac_base: 
                        indice_actual = indice_cac_base

                    indice_cac = indice_actual / indice_cac_base
                    actualizacion = presupuesto_pesos * indice_cac

                    print(f"{presupuesto_pesos} * ({indice_actual} / {indice_cac_base})")
                    print("=", actualizacion)

                    # Restar 
                    actualizacion = actualizacion - monto_pesos
                    print(f" menos {monto_pesos} =", actualizacion)

                    if actualizacion <= 0:
                        resto = -(actualizacion)
                        cac_resto = indice_actual
                        presupuesto.presupuesto_actualizado = 0


                        # Para trasladar a Excel debería sumar el saldo entero.
                        # Debo sumar los presupuestos restantes. 
                        # Y restarles el saldo.
                        suma = 0
                        for p, presu in enumerate (presupuestos_disponibles):

                            monto = presu.presupuesto_actualizado
                            cac__ = presu.indice_cac_base
                                
                            aux__ = indice_actual / cac__
                            monto = monto*aux__

                            #Actualizo Los presupuesos.
                            suma = suma + monto
                                
                        suma = suma - resto 
                        hoja_gastos.range(f"Q{posicion_opy}").value = int(suma)
                #        print("SE SUMAN Ptos RESTANTES SI PAGO>PTO")
                 #       print(f"N°{p} SUMA DE PRESUPUESTOS:", suma)

                        # Si agoté mis presupuestos, acumulo los pagos como saldo.
                        if presupuestos_agotados == len(presupuestos_disponibles):

                 #           print("Presupuestos agotados")
                  #          print("") #CASO N°1

                            #ACUMULO PAGO, EN PESOS Y EN DOLARES, COMO SALDO NEGATIVO
                            acumulado_pesos = 0
                            while posicion_pago < len(pagos_realizados):

                                pago = pagos_realizados[posicion_pago]
                                monto_pesos = pago.monto_pesos_act

                                if not resto == 0:
                                    monto_pesos = monto_pesos + resto
                                    resto = 0

                                posicion_opy = pago.posicion_opy
                                acumulado_pesos = acumulado_pesos + monto_pesos
                  #              print("acumulado_pesos", acumulado_pesos)

                                hoja_gastos.range(f"{col_pesos}{posicion_opy}").value = int(-(acumulado_pesos))
                                posicion_pago +=1
                            return

                        break

                    else:
                        # Almacenar
                        presupuesto.presupuesto_actualizado = actualizacion
                        presupuesto.indice_cac_base = indice_actual

                    # Para trasladar a Excel debería sumar el saldo entero.
                    suma = 0
                    for presu in presupuestos_disponibles:

                        monto = presu.presupuesto_actualizado
                        suma = suma + monto   


                        hoja_gastos.range(f"Q{posicion_opy}").value = int(suma)

                    pago.monto_pesos_act = 0
                    break
                    print("")


    print(f"Actualizado En Pesos: Original: {presu.presupuesto_original}| Actualizado: {presu.presupuesto_actualizado}")
    
    return


# CASO N°3.2 ACTUALIZO EN DOLARES 

def Caso_3_dolares(hoja_gastos, presupuestos_disponibles, pagos_realizados):

    print("ACTUALIZACION EN DOLARES")
    print("")
    
    resto = 0   # afectará si el pago es mayor al presupuesto/saldo

    for n, presupuesto in enumerate(presupuestos_disponibles):

     #   print("Pto. N°", n)
        print(presupuesto.presupuesto_dolares_actualizado)
        print("")

        for nn, pago in enumerate(pagos_realizados):

     #       print("Pago", nn, pago.fecha, pago.monto_dolares)
            
            posicion_opy = pago.posicion_opy
            fecha_pago = pago.fecha
            monto_dolares = pago.monto_dolares
            
            # para evitar pagos ya procesados:
            if monto_dolares == 0:
                continue

            else:
                if not resto == 0:
                    presupuesto_dolares = presupuesto.presupuesto_dolares_actualizado
                    actualizacion = presupuesto_dolares - resto
                    presupuesto.presupuesto_dolares_actualizado = actualizacion
                    resto = 0

                # RESTO PAGO y ALMACENO
                presupuesto_dolares = presupuesto.presupuesto_dolares_actualizado
                actualizacion = presupuesto_dolares - monto_dolares
                print(f" presupuesto dolares {presupuesto_dolares} =")
                print(f"menos {monto_dolares}")
                print("=", actualizacion)
                print("")

                if actualizacion <= 0: 
                    resto = -(actualizacion)
                    presupuesto.presupuesto_dolares_actualizado = 0
                    print("RESTO", resto)

                # Almacenar
                presupuesto.presupuesto_dolares_actualizado = actualizacion
                
                # Para trasladar a Excel debería sumar el saldo entero.
                suma_dolares = 0
                for presu in presupuestos_disponibles:

                    monto_dolares = presu.presupuesto_dolares_actualizado 
                    suma_dolares = suma_dolares + monto_dolares

                    hoja_gastos.range(f"R{posicion_opy}").value = int(suma_dolares)

                pago.monto_dolares = 0
                print("")
            
            
    print("Actualizado En Dolares")
    
    return


# In[ ]:


# 14.2.a - CASO N°3 - EN DOLARES 

# CASO N°3.2 ACTUALIZO 

def Caso_3_dolares(hoja_gastos, presupuestos_disponibles, pagos_realizados):

    
    col_pesos = "Q"
    col_dolares = "R"
    print("")
    print("ACTUALIZA EN DÓLARES")
    print("")


    resto = 0   # afectará si el pago es mayor al presupuesto/saldo
    
    # MIENTRAS TENGA PAGOS
    for posicion_pago, pago in enumerate(pagos_realizados):
        
    #Si el pago tiene subrubro, uso solo los presupuestos con el mismo subrubro
        sub_rubro = (pago.subrubro)
        if not (sub_rubro == "nan") or (sub_rubro == None) or (sub_rubro == np.nan):
        
            presupuestos_con_subrubro = []
            for presu in presupuestos_disponibles:
                
                subr_presu = (presu.subrubro)
                if not (subr_presu == "nan") or not (subr_presu == None) or not(subr_presu == np.nan) and sub_rubro == subr_presu:
                    presupuestos_con_subrubro.append(presu)
                    
            presupuestos_disponibles = presupuestos_con_subrubro
    
        
        # Si el pago ya fue analizado:
        if pago.monto_dolares_act <= 0:
   #         print("continuamos")
            continue
        
        # llevo cuenta de los presupuestos usados
        presupuestos_agotados = sum(1 for presupuesto in presupuestos_disponibles if presupuesto.presupuesto_dolares_actualizado <= 0)
            
        
        for posicion_presu, presupuesto in enumerate(presupuestos_disponibles):

            # Si agoté mis presupuestos, acumulo los pagos como saldo.
            if presupuestos_agotados == len(presupuestos_disponibles):
                
     #           print("Presupuestos agotados")
     #           print("") #CASO N°1
                
                #ACUMULO PAGO, EN _dolares Y EN DOLARES, COMO SALDO NEGATIVO
                acumulado_dolares = 0
                while posicion_pago < len(pagos_realizados):
                    
                    pago = pagos_realizados[posicion_pago]
                    monto_dolares = pago.monto_dolares_act
                    
                    if not resto == 0:
                        monto_dolares = monto_dolares + resto
                        resto = 0
                    
                    posicion_opy = pago.posicion_opy
                    acumulado_dolares = acumulado_dolares + monto_dolares
                    print("acumulado_dolares", acumulado_dolares)
                    
                    hoja_gastos.range(f"{col_dolares}{posicion_opy}").value = int(-(acumulado_dolares))
         #           print(f"col_dolares {col_dolares}, posicion_opy {posicion_opy}")
                    
                    posicion_pago +=1
                return
            
            
            # Si el presupuesto ya fue analizado
            if presupuesto.presupuesto_dolares_actualizado <= 0:
                continue
                            
            else:

        #        print("Pto. N°", posicion_presu)
        #        print(presupuesto.presupuesto_dolares_actualizado)
        #        print("")


        #        print("Pago", posicion_pago, pago.monto_dolares)

                fecha_pago = pago.fecha
                monto_dolares = pago.monto_dolares_act
                indice_actual = pago.base_cac
                posicion_opy = pago.posicion_opy

                if not resto == 0:

                    print("HAY RESTO", resto)
                    print("")
                    presupuesto_dolares = presupuesto.presupuesto_dolares_actualizado
                    print("presupuesto_dolares", presupuesto_dolares)
                    actualizacion = presupuesto_dolares - resto
                    print(f"actualizacion = {presupuesto_dolares} - {resto}")
                    print("=", actualizacion)
                        

                    if actualizacion <= 0:
                        
                        resto = -(actualizacion)
                        presupuesto.presupuesto_dolares_actualizado = 0
                        
                        # Para trasladar a Excel debería sumar el saldo entero.
                        suma = 0
                        for presu in presupuestos_disponibles:
                            monto = presu.presupuesto_dolares_actualizado
                            suma = suma + monto   
                        
                        suma = suma - resto
                        hoja_gastos.range(f"{col_dolares}{posicion_opy}").value = int(suma)
           #             print(f"col_dolares {col_dolares}, posicion_opy {posicion_opy}")
                        

            #            print("SUMA DE PRESUPUESTOS SI PAGO>PTOS ", suma)
                        break
                        
                    else:

                        presupuesto.presupuesto_dolares_actualizado = actualizacion
                        resto = 0
                    

                # ACTUALIZO Y RESTO PAGO 
                presupuesto_dolares = presupuesto.presupuesto_dolares_actualizado

                # Restar 
                actualizacion = presupuesto_dolares - monto_dolares
                print(f" menos {monto_dolares} =", actualizacion)
                
                # Si el pago es mayor al presupuesto:
                if actualizacion <= 0:
                    
                    # se acumula como resto y se actualiza base de dato
                    resto = -(actualizacion)
                    presupuesto.presupuesto_dolares_actualizado = 0
                    
                    # Para trasladar a Excel debería sumar el saldo 
                    # generado por todos los presupuestos (en caso que existan)
                    suma = 0
                    for presu in presupuestos_disponibles:

                        monto = presu.presupuesto_dolares_actualizado
                        suma = suma + monto   

                    suma = suma - resto
                    
                    # Envio saldo a excel
                    hoja_gastos.range(f"{col_dolares}{posicion_opy}").value = int(suma)
        #            print(f"col_dolares {col_dolares}, posicion_opy {posicion_opy}")
       #             print("SUMA DE PTOS SI [PAGO > PTOS] ", suma)
                    
                    # Si agoté mis presupuestos, acumulo los pagos como saldo.
                    if presupuestos_agotados == len(presupuestos_disponibles):

                        print("Presupuestos agotados")
                        print("") #CASO N°1

                        #ACUMULO PAGO, EN dolares Y EN DOLARES, COMO SALDO NEGATIVO
                        acumulado_dolares = 0
                        while posicion_pago < len(pagos_realizados):

                            pago = pagos_realizados[posicion_pago]
                            monto_dolares = pago.monto_dolares_act

                            if not resto == 0:
                                monto_dolares = monto_dolares + resto
                                resto = 0

                            posicion_opy = pago.posicion_opy
                            acumulado_dolares = acumulado_dolares + monto_dolares
                            print("acumulado_dolares", acumulado_dolares)

                            hoja_gastos.range(f"{col_dolares}{posicion_opy}").value = int(-(acumulado_dolares))
                            print(f"col_dolares {col_dolares}, posicion_opy {posicion_opy}")
                            posicion_pago +=1
                        return

                    break

                else:
                    # Almacenar y envío a excel
                    presupuesto.presupuesto_dolares_actualizado = actualizacion
                    hoja_gastos.range(f"{col_dolares}{posicion_opy}").value = int(actualizacion)
        #            print(f"col_dolares {col_dolares}, posicion_opy {posicion_opy}")
                    
               
    print("Actualizado En Dólares")
    
    return


# In[ ]:


# 14.b CASO N°2 - HAY PRESU, NO HAY PAGOS


def caso_2(hoja_gastos,presupuestos_disponibles, pagos_realizados):
    fecha_actual = datetime.now().date()
    fecha_actual = pd.to_datetime(fecha_actual)

#    print("FECHA ACTUAL", fecha_actual)
    fecha_cac_actual = fecha_actual.replace(day=1) - relativedelta(months=1) # FECHA CAC MES ANTERIOR
#     print("FECHA CAC ACTUAL", fecha_cac_actual)


    for presupuesto in presupuestos_disponibles:
#         print("presupuesto",presupuesto)
        monto = presupuesto.presupuesto_actualizado
        indice = presupuesto.indice
        indice_actual = 0
        if indice == "MANO_DE_OBRA" or indice == "MATERIALES" or indice == "COSTO_CONSTRUCCIÓN":
            base = presupuesto.indice_base
            indice_actual = float(cac[indice].loc[cac[indice]["fecha"] == fecha_cac_actual, indice].values[0])   
        else:
            base = presupuesto.indice_cac_base
            indice_actual = float(cac["COSTO_CONSTRUCCIÓN"].loc[cac["COSTO_CONSTRUCCIÓN"]["fecha"] == fecha_cac_actual, "COSTO_CONSTRUCCIÓN"].values[0])   

  #      print("monto", monto)
  #      print("indice", indice)
  #      print("Monto * (indice actual / base)")
  #      print(f"{monto} * {indice_actual} / {base}")
        aux = indice_actual / base
        actualizacion = monto * aux

        #AlOJO EN ESTRUCTURA
        print("PRESUPUESTO ACTUALIZADO =", actualizacion)
        presupuesto.presupuesto_actualizado = actualizacion
        

    return


# In[ ]:


# 14.A CASO N°1 - NO PTO, SI HAY PAGOS

def caso_1(hoja_gastos, presupuestos_disponibles, pagos_realizados):
    

    print("")
    print("CASO N°1: NO HAY PTO. SI HAY PAGOS")
    print("")

    #VARIABLES NECESARIAS DENTRO DE SALDO
    fecha_pago = 0
    acumulado_pesos = 0
    acumulado_dolares = 0

    #ACUMULO PAGO, EN PESOS Y EN DOLARES, COMO SALDO NEGATIVO
    for pago in pagos_realizados:

        fecha_pago = pago.fecha
        monto_pesos = pago.monto_pesos_act
        print("MONTO_PESOS", monto_pesos)
        monto_dolares = pago.monto_dolares_act
        posicion_opy = pago.posicion_opy
        #SUMO
        acumulado_pesos = acumulado_pesos + monto_pesos
        print("acumulado pesos", acumulado_pesos)
        acumulado_dolares = acumulado_dolares + monto_dolares
        print("acumulado dolares", acumulado_dolares)

        hoja_gastos.range(f"Q{posicion_opy}").value = int(-(acumulado_pesos))
        hoja_gastos.range(f"R{posicion_opy}").value = int(-(acumulado_dolares))


    #ALOJO COMO SALDO NEGATIVO DENTRO DE ESTRUCTURA                                
    fecha = fecha_pago
    subrubro = "-"
    saldo_pesos = -(acumulado_pesos)
    saldo_dolares = -(acumulado_dolares)

    clave = fecha
    clave = Saldo(fecha, subrubro, saldo_pesos, saldo_dolares)
    CONTENEDOR[RUBRO][CAT]["saldo"].append(clave)
    
    return


# In[ ]:


# 14.0 - FILTRAR SI HAY SUBRUBRO

def SubRubros(presupuestos_disponibles, pagos_realizados):
                
    # Clasificacion Rubros
    presupuestos_con_sub = []
    presupuestos_sin_sub = []

    # Clasificacion Pagos
    pagos_con_sub = []
    pagos_sin_sub = []
    
    subs_presu = set()
    subs_pagos = set()

    # FILTRAR PRESUPUESTOS
    for presupuesto in presupuestos_disponibles:
        sub = presupuesto.subrubro
        #print(sub)

        if sub is None or (isinstance(sub, float) and np.isnan(sub)) or sub == "nan":
            presupuestos_sin_sub.append(presupuesto)
        else:
            sub_rubro = {f"{sub}": presupuesto}
            presupuestos_con_sub.append(sub_rubro)
            #print(presupuestos_con_sub)
            
            
    if len(presupuestos_con_sub) > 0:
        presupuestos_disponibles = presupuestos_con_sub

        # ORDENAR SUBRUBROS EN LISTA
        subs_presu = set()  # Crear un conjunto vacío para almacenar los nombres únicos

        for diccionario in presupuestos_disponibles:
            for subrubro in diccionario:
                subs_presu.add(subrubro)  # Agregar el nombre del subrubro al conjunto

        # Convertir el conjunto a una lista (si es necesario)
        subs_presu = list(subs_presu)
    
    
    # FILTRAR PAGOS

    for pago in pagos_realizados:
        sub = pago.subrubro
        #print(sub)

        if sub is None or (isinstance(sub, float) and np.isnan(sub)) or sub == "nan":
            pagos_sin_sub.append(pago)
        else:
            sub_rubro = {f"{sub}": pago}
            pagos_con_sub.append(sub_rubro)
            #print(presupuestos_con_sub)


    if len(presupuestos_con_sub) > 0:
        pagos_realizados = pagos_con_sub
        
        subs_pagos = set()  # Crear un conjunto vacío para almacenar los nombres únicos

        for diccionario in pagos_realizados:
            for subrubro in diccionario:
                subs_pagos.add(subrubro)  # Agregar el nombre del subrubro al conjunto

        # Convertir el conjunto a una lista (si es necesario)
        subs_pagos = list(subs_pagos)
    
    
    if subs_pagos or subs_presu:
        filtros = list(set(subs_pagos) | set(subs_presu))
    
    else:
        filtros = None
    
                
    return filtros
                        
    
#if len(subs_pagos) > 0:
#    print(subs_pagos)


# In[ ]:


# 14 - ITERO SOBRE CADA RUBRO Y ACTUALIZO

print("ACTUALIZACION:") 
print("")

col_pesos = "Q"
col_dolares = "R"

gastos_xw = xw.Book(ruta_gastos)
hoja_gastos = gastos_xw.sheets("GASTOS DE OBRA")

for num, RUBRO in enumerate(CONTENEDOR):
    
    # ITERA SOBRE RUBROS
    if num == 0: 
        print("")
        #excepto posicion 0
        
    else:
        print("")
        print("")
        print("RUBRO", RUBRO)
        print("-------------------------------------------------------")
        
        # ITERA SOBRE CATEGORIAS
        for CAT in CONTENEDOR[RUBRO]:
            
            print("---------------- Categoria", CAT)
                      
            # LISTA PRESUPUESTOS Y LISTA PAGOS?
            presupuestos_disponibles = CONTENEDOR[RUBRO][CAT]["presupuesto"]
            pagos_realizados = CONTENEDOR[RUBRO][CAT]["pagos"]
            
            # CHEQUEO SI EL RUBRO TIENE SUB-RUBROS:
            filtros = SubRubros(presupuestos_disponibles, pagos_realizados)
            
            if filtros == None:
                
                print("NO TIENE SUBRUBRO")
                
                # SIN PRESUPUESTO:
                if len(presupuestos_disponibles) == 0:

                    if not len(pagos_realizados) == 0:

                        # CASO N°1 = HAY PAGOS PERO NO PRESUPUESTO
                        # Acumulo pagos como saldo negativo.
                        
                        print("-----CASO N°1 = HAY PAGOS PERO NO PRESUPUESTO")
                        primer_caso = caso_1(hoja_gastos,presupuestos_disponibles, pagos_realizados)    

                    else:
                        print("No hay pagos, ni presupuesto")

                # CON PRESUPUESTO
                else:

                    # HAY PAGOS?
                    if len(pagos_realizados) == 0:

                        # CASO N°2: HAY PRESUPUESTO, NO HAY PAGOS. 
                        # Se actualizan los presupuestos

                        print("")
                        print("-----CASO N°2 HAY PRESUPUESTO Y NO HAY PAGOS")
                        segundo_caso = caso_2(hoja_gastos,presupuestos_disponibles, pagos_realizados)
                        
                    # CASO N°3: HAY PRESUPUESTO/S y PAGO/S.
                    else:

                        print("-----CASO N°3: HAY PTO Y HAY PAGOS")

                        actualizar_pesos = Caso_3_pesos(hoja_gastos,presupuestos_disponibles, pagos_realizados)
                        actualizar_dolares = Caso_3_dolares(hoja_gastos,presupuestos_disponibles, pagos_realizados)
                
            else:
                
                #print("TIENE SUBRUBROS:")

                for filtro in filtros:

                    presupuestos_disponibles = CONTENEDOR[RUBRO][CAT]["presupuesto"]
                    pagos_realizados = CONTENEDOR[RUBRO][CAT]["pagos"]

                    print("FILTRO:", filtro)

                    #FILTRO LOS PRESUPUESTOS SEGUN SUBRUBRO
                    nueva_lista_presu = []
                    for presu in presupuestos_disponibles:
                        if (presu.subrubro) == filtro:
                            nueva_lista_presu.append(presu)
                            
                    #FILTRO LOS PAGOS SEGUN SUBRUBRO
                    nueva_lista_pagos = []
                    for pago in pagos_realizados:
                        if (pago.subrubro) == filtro:
                            nueva_lista_pagos.append(pago)
                            
                    #print("LISTAS NUEVAS")
                    #print("presus", nueva_lista_presu)
                    #print("pagos", nueva_lista_pagos)
                    #print("")

                    presupuestos_disponibles = nueva_lista_presu
                    pagos_realizados = nueva_lista_pagos

                    # SIN PRESUPUESTO:
                    if len(presupuestos_disponibles) == 0:

                        if not len(pagos_realizados) == 0:

                            # CASO N°1 = HAY PAGOS PERO NO PRESUPUESTO
                            # Acumulo pagos como saldo negativo.
                            print("-----CASO N°1 = HAY PAGOS PERO NO PRESUPUESTO")
                            primer_caso = caso_1(hoja_gastos,presupuestos_disponibles, pagos_realizados)    

                        else:
                            print("No hay pagos, ni presupuesto")

                    # CON PRESUPUESTO
                    else:
                        # HAY PAGOS?
                        if len(pagos_realizados) == 0:

                            # CASO N°2: HAY PRESUPUESTO, NO HAY PAGOS. 
                            # Se actualizan los presupuestos

                            print("")
                            print("-----CASO N°2 HAY PRESUPUESTO Y NO HAY PAGOS")
                            segundo_caso = caso_2(hoja_gastos,presupuestos_disponibles, pagos_realizados)

                        # CASO N°3: HAY PRESUPUESTO/S y PAGO/S.
                        else:

                            print("-----CASO N°3: HAY PTO Y HAY PAGOS")

                            actualizar_pesos = Caso_3_pesos(hoja_gastos,presupuestos_disponibles, pagos_realizados)
                            actualizar_dolares = Caso_3_dolares(hoja_gastos,presupuestos_disponibles, pagos_realizados)

                print("")

# GUARDO CON XL-WINGS
gastos_xw.save()


print("FINALIZA ACTUALIZACION")           


# In[ ]:


# 15 - ALMACENO RESULTADOS EN HOJA RESUMEN

def RESUMEN():
    
    # ÚLTIMOS VALORES CAC
    
    mat= (cac["MATERIALES"].loc[1, "MATERIALES"])
    cc = (cac["COSTO_CONSTRUCCIÓN"].loc[1, "COSTO_CONSTRUCCIÓN"])
    man = (cac["MANO_DE_OBRA"].loc[1, "MANO_DE_OBRA"])
    
    resumen_pesos.range("P4").value = float(cc)
    resumen_pesos.range("Q4").value = float(mat)
    resumen_pesos.range("R4").value = float(man)
    
    # ÚLTIMOS VALORES DÓLAR
    
    # BLUE
    a = 0
    blu = (dolar["BLUE"].loc[a, "blue"])
    while blu == None or np.isnan(blu) or blu == "nan" or blu == 0:
        a += 1
        blu = (dolar["BLUE"].loc[a, "blue"])
        
    
    # MEP
    b = 0
    me = dolar["MEP"].loc[b, "mep"]
    while me == None or me == "nan" or np.isnan(me) or me == 0:
        b += 1
        me = (dolar["MEP"].loc[b, "mep"])
        
    
    # OFICIAL
    c = 0
    ofi = (dolar["OFICIAL"].loc[c, "oficial"])
    while ofi == None or ofi == "nan" or ofi == 0 or np.isnan(ofi):
        c += 1
        ofi = (dolar["OFICIAL"].loc[c, "oficial"])
    
    resumen_pesos.range("S4").value = float(blu)
    resumen_pesos.range("T4").value = float(me)
    resumen_pesos.range("U4").value = float(ofi)
    
    print(f" Valores dolar: Blue:{blu} | MEP:{me} | Oficial:{ofi}")
    
    
    
    presupuesto_pesos = pd.read_excel(ruta_presupuesto, 
                                      sheet_name = "RESUMEN DE OBRA PESOS",
                                      usecols = ("ÍTEM", "RUBRO", "CATEGORÍA"),
                                      skiprows = 3,
                                      )
    rubro = 0 #Inicializo 
    index_open = 4
    
    for posicion in (presupuesto_pesos.index):
        index_open +=1
        lector_rubro = presupuesto_pesos.loc[presupuesto_pesos.index == posicion, "ÍTEM"].values
        
        try: 
            int(lector_rubro)
            rubro = int(lector_rubro)
        except ValueError:
            rubro = rubro
            
    
        # PONGO EN CERO LOS RESULTADOS ANTERIORES PARA EVITAR CONFUCIONES
         # En Pesos
        resumen_pesos.range(f"D{index_open}").value = None
        resumen_pesos.range(f"G{index_open}").value = None
        resumen_pesos.range(f"I{index_open}").value = None
         
         # En Dolares
        resumen_dolares.range(f"D{index_open}").value = None
        resumen_dolares.range(f"G{index_open}").value = None
        resumen_dolares.range(f"H{index_open}").value = None
        resumen_dolares.range(f"I{index_open}").value = None
        resumen_dolares.range(f"K{index_open}").value = None


            
        
        categoria = str(presupuesto_pesos.loc[presupuesto_pesos.index == posicion, "CATEGORÍA"].values[0]).upper().replace(" ", "_")
        
        
        
        
        # PRESUPUESTOS
        
        presupuestos_disponibles = CONTENEDOR[rubro][categoria]["presupuesto"]
        pagos_realizados = CONTENEDOR[rubro][categoria]["pagos"]
        
        suma_original = 0 
        suma_actualizado = 0
        suma_dolares_original = 0
        suma_mep =0
        suma_blue =0
        suma_dolares_actualizado = 0

        if len(presupuestos_disponibles) > 0:
            for presupuesto in presupuestos_disponibles:

                # SUMA PRESUPUESTO ORIGINAL
                monto_original = presupuesto.presupuesto_original
                suma_original = suma_original + monto_original

                # SUMA DE PRESUPUESTOS ACTUALIZADOS = MONTO ACTUAL                
                monto_actualizado = presupuesto.presupuesto_actualizado
                suma_actualizado = suma_actualizado + monto_actualizado

                # SUMA PRESUPUESTO ORIGINAL EN USD
                monto_dolares_original = presupuesto.presupuesto_dolares
                suma_dolares_original = suma_dolares_original + monto_dolares_original

                # SUMA PRESUPUESTO ORIGINAL EN USD
                monto_dolares_actualizado = presupuesto.presupuesto_dolares_actualizado
                suma_dolares_actualizado = suma_dolares_actualizado + monto_dolares_actualizado

                # SUMA PRESUPUESTO MEP
                monto_mep = presupuesto.presupuesto_mep
                suma_mep = suma_mep + monto_mep

                # SUMA PRESUPUESTO BLUE
                monto_blue = presupuesto.presupuesto_blue
                suma_blue = suma_blue + monto_blue


            #ENVIO A EXCEL CON XLWINGS
            #print("presupuestos:")
            #print("")
            # En pesos
            resumen_pesos.range(f"D{index_open}").value = int(suma_original)
            resumen_pesos.range(f"I{index_open}").value = int(suma_actualizado)


            # En dolares
            resumen_dolares.range(f"D{index_open}").value = int(suma_dolares_original)
            resumen_dolares.range(f"G{index_open}").value = int(suma_mep)
            resumen_dolares.range(f"H{index_open}").value = int(suma_blue)
            resumen_dolares.range(f"K{index_open}").value = int(suma_dolares_actualizado)




            # PAGOS
            # print("")
            # print("pagos")


        suma_original = 0
        suma_dolares = 0
        if len(pagos_realizados) > 0:

            for pago in pagos_realizados:

                monto_original = pago.monto_pesos
                suma_original = suma_original + monto_original

                monto_dolares = pago.monto_dolares
                suma_dolares = suma_dolares + monto_dolares


        #ENVIO A EXCEL CON XLWINGS
        resumen_pesos.range(f"G{index_open}").value = int(suma_original)

        # U$D
        resumen_dolares.range(f"I{index_open}").value = int(suma_dolares)

        #print("En pesos: ", suma_original, "En dolares", suma_dolares)
        #print("")

            
    # GUARDO CON OPENPYXL
    
    presupuesto_xw.save(ruta_presupuesto)  
    
    return           
                       
resumen = RESUMEN()



# In[ ]:


# 16 AVISO DE CIERRE

ventana = tk.Tk()
ventana.withdraw()  # Ocultar la ventana principal
ventana.configure(bg="black")

# Mostrar el mensaje de error en una ventana emergente
messagebox.showinfo("Mensaje", "Actualización Finalizada")


ventana.destroy()  # Cerrar la ventana emergente


# In[ ]:


input("El programa ha finalizado. Presione 'Enter' para cerrar terminal")

